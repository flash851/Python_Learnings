#Deleting a sheet from a xml file

import openpyxl
# workbook=openpyxl.load_workbook("E:\Employees.xlsx")
# sheet=workbook['Testsheet1']     
# workbook.remove(sheet)
# workbook.save('E:\Employees.xlsx')

workbook=openpyxl.load_workbook("E:\Employees.xlsx")
sheet=workbook['EmployeeData']

#gives sheet title
print(sheet.title)

#available function
print(dir(sheet))

# getting active cell
print(sheet.active_cell)

# getting dimensions
print(sheet.dimensions)

print(sheet.sheet_format)

print(sheet.sheet_properties)

print(sheet.sheet_state)

print(sheet.sheet_view)

#print rows and columns count
print(sheet.max_row)
print(sheet.max_column)

# iterating the sheet rows
for i in sheet.values:
	print(i)