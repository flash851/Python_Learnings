# importing package
import openpyxl
print("anuj")

# Craeting workbook
workbook=openpyxl.load_workbook("E:\Employees.xlsx")

# getting properties of xsls file
print(workbook.properties)

# Getting the sheetnames
print(workbook.sheetnames)

# Creting a new sheet
workbook.create_sheet('Testsheet')

# Saving a sheet
workbook.save('E:\Employees.xlsx')

# removing a sheet
sheet=workbook['Testsheet']     
workbook.remove(sheet)

#del workbook['Testsheet']

# Saving a sheet
workbook.save('E:\Employees.xlsx')